import openpyxl
import pandas as pd
from aiogram import executor, types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import StatesGroup, State
from aiogram.types import ReplyKeyboardMarkup
from openpyxl.reader.excel import load_workbook

from functions.uzpharmExcell import Uz_pharm_excell, Uz_pharm_excell2
from keys import dp, bot, ADMIN_ID
from functions.uzpharmParse import uz_pharm_parse
from milliycatalog.getlist import getlistall
import asyncio


mainBtn = ReplyKeyboardMarkup(resize_keyboard=True)
mainBtn.add("Send Excell", "UzPharm-Control")
mainBtn.add("MilliyCatalog")


class Form(StatesGroup):
    sendExcell = State()
    uzpharm = State()


@dp.message_handler(commands='start', user_id=ADMIN_ID)
async def start(msg: types.Message):
    await msg.answer("Assalomu alaykum, xush kelibsiz, bot ishlayapti", reply_markup=mainBtn)


@dp.message_handler(text="MilliyCatalog", user_id=ADMIN_ID)
async def start(msg: types.Message):
    await msg.answer(text="bu jarayot bir necha soatni o'z ichiga olishi mumkin, "
                          "tayyor bo'lganda sizga excell yuboramiz",
                     reply_markup=mainBtn)

    loop = asyncio.get_running_loop()
    loop.run_in_executor(None, lambda: asyncio.run(getlistall(msg.from_user.id)))

###################################################################################################


@dp.message_handler(text="Send Excell", user_id=ADMIN_ID)
async def sendDoc(message: types.Message):
    backBtn = ReplyKeyboardMarkup(resize_keyboard=True)
    backBtn.add("🔙Back")
    await message.answer("Excellni yuboring", reply_markup=backBtn)
    await Form.sendExcell.set()


@dp.message_handler(state='*', text="🔙Back", user_id=ADMIN_ID)
async def sendDoc(message: types.Message, state: FSMContext):
    try:
        await state.finish()
    except:
        pass
    await message.answer("Bosh menyu", reply_markup=mainBtn)


@dp.message_handler(state='*', content_types="document", user_id=ADMIN_ID)
async def send(message: types.Message, state: FSMContext):
    try:
        with open(r"main_name.txt", 'r') as f:
            main_name = f.read()
        if not main_name:
            main_name = "Реестр референтных"
    except:
        main_name = "Реестр референтных"

    msg0 = await message.answer("qabul qildim")
    try:
        if message.document.mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" \
                and main_name in message.document.file_name:
            await msg0.edit_text("Jarayon boshlandi")
            fileN = message.document.file_name
            fileID = message.document.file_id
            res = await bot.get_file(fileID)
            await bot.download_file(res.file_path, f'fileDown/{fileN}')

            file_cvs = fileN[:-5]
            msg1 = await message.answer("excell tayyorlanmoqcda ...")
            await editExcel(file_name=fileN, file_cvs=file_cvs)
            await msg1.edit_text("Bazaga kiritish boshlandi")
            await msg1.edit_text(f"Bazaga kiritilyapdi")
            insert_row = await Uz_pharm_excell(msg=msg1,  file_name=fileN)
            df = pd.read_excel(f"fileDown/{fileN}", engine='openpyxl', dtype=object, header=None)
            await msg1.edit_text(f"Bazaga kiritildi\n\nBazaga {len(df.values.tolist())}tadan {insert_row}tasi kiritildi")

            fileLoc = f"fileDown/{file_cvs}.csv"
            await message.answer_document(document=open(fileLoc, 'rb'))
            await message.answer("tugallandi")
        try:
            await state.finish()
        except Exception as ex:
            print(ex)
            pass
    except Exception as ex:
        print(ex)
        await message.answer(f"Error: {str(ex)}")


async def editExcel(file_name, file_cvs):
    url = f'fileDown/{file_name}'
    wb_obj = openpyxl.load_workbook(url)
    sheet_obj = wb_obj.active
    nums = sheet_obj.max_row + 1

    workbook = load_workbook(filename=url)
    worksheet = workbook.active
    day = file_cvs[-10:-8]
    month = file_cvs[-7:-5]
    year = file_cvs[-4:]
    data = year+'-'+month+'-'+day
    col_num = worksheet.max_column + 1
    for row_num in range(1, nums):
        if row_num == 1:
            worksheet.cell(row=row_num, column=col_num, value="Date")
        else:
            worksheet.cell(row=row_num, column=col_num, value=data)
    workbook.save(filename=url)

    read_file = pd.read_excel(f'fileDown/{file_name}')
    read_file.to_csv(f'fileDown/{file_cvs}.csv', header=True, index=False)
    pd.DataFrame(pd.read_csv(f'fileDown/{file_cvs}.csv'))

###################################################################################################


@dp.message_handler(text="UzPharm-Control", user_id=ADMIN_ID)
async def pharm(message: types.Message):
    await message.answer("qabul qildim")
    try:
        await message.answer("Jarayon boshlandi")
        await uz_pharm_parse()
        read_file = pd.read_excel("uzpharmDown/uzpharm.xlsx")
        read_file.to_csv(f'uzpharmDown/uzpharm.csv', header=True, index=False)
        pd.DataFrame(pd.read_csv(f'uzpharmDown/uzpharm.csv'))
        await message.answer_document(document=open("uzpharmDown/uzpharm.csv", 'rb'))
        await message.answer_document(document=open("uzpharmDown/uzpharm.xlsx", 'rb'))

        excellFile = "uzpharmDown/uzpharm.xlsx"
        df = pd.read_excel("uzpharmDown/uzpharm.xlsx", engine='openpyxl', dtype=object, header=None)

        user_id = message.from_user.id
        if str(user_id) in ADMIN_ID:
            await message.answer("bazaga kiritish boshlandi")
            insert_row = await Uz_pharm_excell2(excellFile)
            await message.answer(f"Bazaga kiritish tugallandi\n\nBazaga {len(df.values.tolist())}ta ma'lumotdan {insert_row}tasi kiritildi")
    except Exception as ex:
        await message.answer(f"Error: {str(ex)}")


@dp.message_handler(content_types=types.ContentType.TEXT, user_id=ADMIN_ID)
async def change_name(msg: types.Message):
    with open('main_name.txt', 'w') as f:
        f.write(msg.text)
    await msg.answer("nom o'zgartirildi")


if __name__ == '__main__':
    executor.start_polling(dp)
